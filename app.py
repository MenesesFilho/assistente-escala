
from flask import Flask, request, send_file, render_template, abort
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from io import BytesIO

app = Flask(__name__)

SEMANA_DIAS = ['segunda-feira','terça-feira','quarta-feira','quinta-feira','sexta-feira','sábado','domingo']
DIAS_MAP_EN2PT = {'Monday':'segunda-feira','Tuesday':'terça-feira','Wednesday':'quarta-feira',
                  'Thursday':'quinta-feira','Friday':'sexta-feira','Saturday':'sábado','Sunday':'domingo'}

def parse_hour_start(s):
    try:
        h = str(s).split(' ')[0]
        return int(h.split(':')[0])
    except Exception:
        return np.nan

def carregar_vendas_dia(f):
    try:
        df = pd.read_excel(f, sheet_name=0, skiprows=3)
    except:
        f.seek(0)
        df = pd.read_excel(f, sheet_name=0)
    cols = df.columns
    # mapeia nomes tolerantes
    nomes = {'Data':'Data','BALCÃO':'BALCÃO','DELIVERY':'DELIVERY','Total':'Total'}
    m = {c:c for c in cols}
    for k in list(nomes.keys()):
        cand = [c for c in cols if str(c).strip().lower()==k.lower()]
        if cand: m[nomes[k]] = cand[0]
    # colunas usadas
    use_cols = [m.get('Data'), m.get('BALCÃO'), m.get('DELIVERY'), m.get('Total')]
    use_cols = [c for c in use_cols if c in df.columns]
    df = df[use_cols].dropna(how='all').copy()
    if m.get('Data') in df.columns:
        df['Data'] = pd.to_datetime(df[m['Data']].astype(str).str[:10], errors='coerce', dayfirst=True)
        df['DiaSemana'] = df['Data'].dt.day_name().map(DIAS_MAP_EN2PT)
    else:
        if 'DiaSemana' not in df.columns:
            raise ValueError("Planilha de vendas por dia precisa ter coluna Data ou DiaSemana.")
    total_col = m.get('Total')
    if total_col not in df.columns:
        poss = [c for c in ['BALCÃO','DELIVERY'] if c in m and m[c] in df.columns]
        if poss:
            df['Total_calc'] = df[m['BALCÃO']].fillna(0) + df[m['DELIVERY']].fillna(0)
            total_col = 'Total_calc'
        else:
            raise ValueError("Não encontrei a coluna 'Total' nem BALCÃO/DELIVERY para somar.")
    day_weights = df.groupby('DiaSemana')[total_col].mean()
    day_weights = (day_weights / day_weights.sum()).reindex(SEMANA_DIAS).fillna(1/7)
    return day_weights

def carregar_vendas_hora(f, abertura_h, fechamento_h):
    try:
        df = pd.read_excel(f, sheet_name=0, skiprows=3)
    except:
        f.seek(0)
        df = pd.read_excel(f, sheet_name=0)
    intervalo_col = 'Intervalo' if 'Intervalo' in df.columns else df.columns[0]
    use_col = None
    for c in ['Vendas','% Fat.','%Fat.','Percentual','Qtde']:
        if c in df.columns and not df[c].isna().all():
            use_col = c; break
    if use_col is None:
        df['VendasUse'] = 1
    else:
        df['VendasUse'] = df[use_col]
    df['HoraInicial'] = df[intervalo_col].apply(parse_hour_start)
    df = df.dropna(subset=['HoraInicial']).copy()
    df['HoraInicial'] = df['HoraInicial'].astype(int)
    oper_hours = list(range(abertura_h-1, fechamento_h+1))  # inclui 1h antes e a última hora operando
    df = df[df['HoraInicial'].isin(oper_hours)]
    hour_weights = df.groupby('HoraInicial')['VendasUse'].sum()
    if hour_weights.sum() == 0 or hour_weights.isna().all():
        hour_weights = pd.Series({h:1 for h in oper_hours})
    hour_weights = hour_weights / hour_weights.sum()
    return hour_weights

def gerar_escala(modelo_bytes, vendas_dia_bytes, vendas_hora_bytes,
                 loja, num_func, abertura, fechamento, tipo_escala, carga):
    # pesos
    day_w = carregar_vendas_dia(vendas_dia_bytes)
    vendas_dia_bytes.seek(0)
    hour_w = carregar_vendas_hora(vendas_hora_bytes, abertura, fechamento)
    vendas_hora_bytes.seek(0)

    hours = sorted(hour_w.index.tolist())  # inclui (abertura-1) .. (fechamento)
    # Definições de cobertura base (regras)
    base_coverage = 2  # em geral 2 pessoas na frente
    # hora de menor peso do dia pode cair para 1
    h_min = hour_w.idxmin()
    # garantir 1h antes da abertura e 1h após fechamento
    pre_open = abertura - 1
    post_close = fechamento  # cobertura até fechamento+1 é parte do turno, mas "hora útil" final do atendimento é fechamento

    # Demanda por slot hora/dia
    needs = {d: {h: base_coverage for h in hours} for d in SEMANA_DIAS}
    for d in SEMANA_DIAS:
        # menor hora do dia com 1 pessoa
        if h_min in needs[d]:
            needs[d][h_min] = 1
        # 1 pessoa na hora pré-abertura
        if pre_open in needs[d]:
            needs[d][pre_open] = max(needs[d][pre_open], 1)
        # mínimo 2 no fechamento
        if post_close in needs[d]:
            needs[d][post_close] = max(needs[d][post_close], 2)

    # Funcionários e folgas 5x2 evitando sábados/domigos
    employees = [f"Funcionário {i}" for i in range(1, num_func+1)]
    weekday_pairs = [("segunda-feira","terça-feira"),("terça-feira","quarta-feira"),
                     ("quarta-feira","quinta-feira"),("quinta-feira","sexta-feira")]
    folgas = {emp: weekday_pairs[(i % len(weekday_pairs))] for i, emp in enumerate(employees)}

    # Define papéis fixos (abertura/meio/fechamento) por funcionário
    # Proporção baseada no pico próximo do fechamento
    close_peak = hour_w.loc[[h for h in hours if h >= abertura+4 and h <= fechamento]].sum()
    open_peak  = hour_w.loc[[h for h in hours if h <= abertura+1]].sum()
    # regra simples: pelo menos 2 closers, 1 opener. o resto vai para meio
    n_close = max(2, int(round(num_func * float(close_peak))))
    n_open  = max(1, int(round(num_func * float(open_peak))))
    n_mid   = max(0, num_func - n_close - n_open)
    if n_mid < 0:
        n_mid = 0

    roles = {}
    for idx, emp in enumerate(employees):
        if idx < n_open:
            roles[emp] = "abertura"
        elif idx < n_open + n_mid:
            roles[emp] = "meio"
        else:
            roles[emp] = "fechamento"

    # Turnos (8h padrão) + variações 9h para atingir 44h semanais
    #  Abertura padrão 8h: (A-1 .. A+3) / (A+4 .. A+8)
    st_open_8 = (abertura-1, abertura+3, abertura+4, abertura+8)
    #  Abertura 9h (+1h no segundo bloco): (A-1 .. A+3) / (A+4 .. A+9)
    st_open_9 = (abertura-1, abertura+3, abertura+4, abertura+9)

    #  Meio 8h: (A+1 .. A+5) / (A+6 .. A+10)
    st_mid_8 = (abertura+1, abertura+5, abertura+6, abertura+10)
    #  Meio 9h (+1h no segundo bloco): (A+1 .. A+5) / (A+6 .. A+11)
    st_mid_9 = (abertura+1, abertura+5, abertura+6, abertura+11)

    #  Fechamento 8h: (A+4 .. A+8) / (A+9 .. F+1)
    st_close_8 = (abertura+4, abertura+8, abertura+9, fechamento+1)
    #  Fechamento 9h: começa 1h antes (+1h no primeiro bloco): (A+3 .. A+8) / (A+9 .. F+1)
    st_close_9 = (abertura+3, abertura+8, abertura+9, fechamento+1)

    def shift_hours(st): s1,e1,s2,e2 = st; return (e1-s1)+(e2-s2)

    # Construção do cronograma por funcionário/dia
    schedule = {emp: {d: "Folga" for d in SEMANA_DIAS} for emp in employees}
    emp_hours = {emp: 0 for emp in employees}
    emp_days = {emp: 0 for emp in employees}
    MAX_HORAS_DIA = 10
    MAX_DIAS = 5
    TARGET_SEMANA = 44  # úteis

    # Ordem de dias por peso (para escolher 9h nos dias mais fortes)
    dias_por_peso = sorted(SEMANA_DIAS, key=lambda d: float(day_w.loc[d]), reverse=True)

    # Aplicar turnos fixos por papel, distribuindo 4 dias de 9h + 1 dia de 8h
    for emp in employees:
        papel = roles[emp]
        # quais turnos usar
        if papel == "abertura":
            st8, st9 = st_open_8, st_open_9
        elif papel == "meio":
            st8, st9 = st_mid_8, st_mid_9
        else:
            st8, st9 = st_close_8, st_close_9

        # folgas predefinidas
        offs = folgas[emp]
        for d in offs:
            schedule[emp][d] = "Folga"

        # Seleciona 5 dias de trabalho (evita folgas) em ordem de peso do dia
        dias_trabalho = [d for d in dias_por_peso if schedule[emp][d] == "Folga"]
        # acima pegou só folga; precisamos os outros:
        dias_trabalho = [d for d in dias_por_peso if schedule[emp][d] != "Folga"] if dias_trabalho == dias_por_peso else [d for d in dias_por_peso if schedule[emp][d] != "Folga"]
        # Vamos construir explicitamente: 5 primeiros dias do ranking que não são folga
        dias_trabalho = []
        for d in dias_por_peso:
            if schedule[emp][d] == "Folga":
                continue
            dias_trabalho.append(d)
            if len(dias_trabalho) == MAX_DIAS:
                break

        # quatro dias de 9h nos dias mais pesados e um dia de 8h no mais leve dentre os 5 escolhidos
        if len(dias_trabalho) < MAX_DIAS:
            # complemento com qualquer dia que não seja folga
            for d in SEMANA_DIAS:
                if schedule[emp][d] != "Folga" and d not in dias_trabalho:
                    dias_trabalho.append(d)
                if len(dias_trabalho) == MAX_DIAS:
                    break

        dias_9h = dias_trabalho[:4]
        dia_8h = dias_trabalho[4] if len(dias_trabalho) >= 5 else (dias_trabalho[-1] if dias_trabalho else None)

        # aplica
        for d in dias_trabalho:
            st = st9 if d in dias_9h else st8
            dh = shift_hours(st)
            if dh > MAX_HORAS_DIA:  # segurança
                st = st8; dh = shift_hours(st8)
            schedule[emp][d] = f"{st[0]:02d}:00 - {st[1]:02d}:00 / {st[2]:02d}:00 - {st[3]:02d}:00"
            emp_hours[emp] += dh
            emp_days[emp] += 1

        # Ajuste fino: se não chegou a 44h, tenta promover um dia 8h para 9h (se possível)
        while emp_hours[emp] < TARGET_SEMANA and emp_days[emp] <= MAX_DIAS:
            for d in dias_trabalho:
                # promove para 9h se não for 9h
                if papel == "abertura":
                    st = st_open_9
                elif papel == "meio":
                    st = st_mid_9
                else:
                    st = st_close_9
                # se já está 9h, pula
                atual = schedule[emp][d]
                if atual.lower() == "folga":
                    continue
                # checa se já é 9h
                p1, p2 = atual.split('/')
                s1,e1 = [x.strip() for x in p1.split('-')]
                s2,e2 = [x.strip() for x in p2.split('-')]
                dh_atual = (int(e1[:2])-int(s1[:2]))+(int(e2[:2])-int(s2[:2]))
                if dh_atual >= 9:
                    continue
                novo = f"{st[0]:02d}:00 - {st[1]:02d}:00 / {st[2]:02d}:00 - {st[3]:02d}:00"
                schedule[emp][d] = novo
                emp_hours[emp] += (9 - dh_atual)
                if emp_hours[emp] >= TARGET_SEMANA:
                    break
            else:
                break  # não encontrou nada para promover

    # Preencher planilha
    wb = load_workbook(modelo_bytes)

    # ESCALA SEMANAL B2:H13 (funcionários nas linhas, seg..dom nas colunas)
    ws_sem = wb['ESCALA SEMANAL']
    for i, emp in enumerate(employees):
        for j, d in enumerate(SEMANA_DIAS):
            val = schedule[emp][d]
            ws_sem.cell(row=2+i, column=2+j, value=val)

    # Escala San Paolo E18:AI65 (4 linhas por funcionário)
    ws_sp = wb['Escala San Paolo']
    col_start = 5  # E
    row_start = 18
    for idx, emp in enumerate(employees):
        base_row = row_start + idx*4
        for j, d in enumerate(SEMANA_DIAS):
            val = schedule[emp][d]
            if not val or str(val).lower() == "folga":
                for r in range(4): ws_sp.cell(row=base_row+r, column=col_start+j, value="Folga")
            else:
                p1, p2 = val.split("/")
                s1,e1 = [x.strip() for x in p1.split("-")]
                s2,e2 = [x.strip() for x in p2.split("-")]
                ws_sp.cell(row=base_row+0, column=col_start+j, value=s1)
                ws_sp.cell(row=base_row+1, column=col_start+j, value=e1)
                ws_sp.cell(row=base_row+2, column=col_start+j, value=s2)
                ws_sp.cell(row=base_row+3, column=col_start+j, value=e2)
        if ws_sp.cell(row=base_row, column=2).value in (None, ""):
            ws_sp.cell(row=base_row, column=2, value=emp)
        if ws_sp.cell(row=base_row, column=3).value in (None, ""):
            ws_sp.cell(row=base_row, column=3, value=roles.get(emp,"Atendente").capitalize())

    # Funcionários por Hora (T3) B2:H21
    ws_t3 = wb['Funcionários por Hora (T3)']
    hours_full = list(range(abertura-1, fechamento+1))  # sem a hora de limpeza extra (fechamento+1) na visualização T3
    count = pd.DataFrame(0, index=hours_full, columns=SEMANA_DIAS)
    def hour_to_int(hhmm): return int(hhmm.split(':')[0])
    for emp in employees:
        for d in SEMANA_DIAS:
            val = schedule[emp][d]
            if val and str(val).lower() != "folga":
                try:
                    p1, p2 = val.split('/')
                    s1,e1 = [x.strip() for x in p1.split('-')]
                    s2,e2 = [x.strip() for x in p2.split('-')]
                    for hh in range(hour_to_int(s1), hour_to_int(e1)):  # bloco 1
                        if hh in count.index: count.loc[hh, d] += 1
                    for hh in range(hour_to_int(s2), hour_to_int(e2)):  # bloco 2
                        if hh in count.index: count.loc[hh, d] += 1
                except: 
                    pass
    for i, hh in enumerate(hours_full):
        for j, d in enumerate(SEMANA_DIAS):
            ws_t3.cell(row=2+i, column=2+j, value=int(count.loc[hh, d]))

    # INFORMAÇÕES
    ws_info = wb['INFORMAÇÕES']
    info_map = {
        "Loja": loja,
        "Abertura": f"{abertura:02d}:00",
        "Fechamento": f"{fechamento:02d}:00",
        "Funcionários": str(num_func),
        "Tipo de escala": tipo_escala,
        "Carga horária semanal": f"{carga}h úteis",
        "Observações": "Regras aplicadas: 1h pré-abertura; 1h pós-fechamento; min 2 em operação (1 no menor pico); turnos fixos por papel; 44h/semana (4x9h + 1x8h); folgas 5x2 seg-sex; min 2 no fechamento."
    }
    for row in ws_info.iter_rows(min_row=1, max_row=ws_info.max_row, min_col=1, max_col=3):
        label = str(row[0].value) if row[0].value is not None else ""
        if label in info_map:
            row[1].value = info_map[label]

    out = BytesIO()
    wb.save(out); out.seek(0)
    return out

@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")

@app.route("/gerar", methods=["POST"])
def gerar():
    try:
        loja = request.form.get("loja", "San Paolo")
        num_func = int(request.form.get("num_func", "10"))
        abertura = int(request.form.get("abertura", "10:00").split(":")[0])
        fechamento = int(request.form.get("fechamento", "22:00").split(":")[0])
        tipo_escala = request.form.get("tipo_escala", "5x2")
        carga = int(request.form.get("carga", "44"))

        modelo = request.files.get("modelo")
        vendas_dia = request.files.get("vendas_dia")
        vendas_hora = request.files.get("vendas_hora")
        if not modelo or not vendas_dia or not vendas_hora:
            return abort(400, "Envie os três arquivos (modelo, vendas por dia e vendas por hora).")

        out = gerar_escala(modelo, vendas_dia, vendas_hora, loja, num_func, abertura, fechamento, tipo_escala, carga)
        return send_file(out, as_attachment=True, download_name="ESCALA_GERADA.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return abort(500, f"Erro ao gerar escala: {e}")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
